import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.toast import ToastNotification
from ttkbootstrap.validation import add_regex_validation
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook



class Hospital(ttk.Frame):
    def __init__(self, master_window):
        super().__init__(master_window, padding=(40, 30))
        self.pack(fill=BOTH, expand=YES)
        self.name = ttk.StringVar(value="")
        self.address = ttk.StringVar(value="")
        self.doctor_handling = ttk.StringVar(value="")
        self.age = ttk.DoubleVar(value=0)
        self.data_df = pd.DataFrame(columns=["Name", "Address", "Doctor Handling", "Age", "Queue Number"])
        self.colors = master_window.style.colors
        self.queue_counter = 0  # Counter to track the queue number

        instruction_text = "Masukkan data dengan tepat: "
        instruction = ttk.Label(self, text=instruction_text, width=50)
        instruction.pack(fill=X, pady=10)

        self.create_form_entry("Full Name: ", self.name)
        self.create_form_entry("Address: ", self.address)
        self.create_form_entry("Doctor Handling: ", self.doctor_handling)
        self.final_score_input = self.create_form_entry("Age: ", self.age)
        self.create_buttonbox()

        # Center the window on the screen
        master_window.update_idletasks()
        width = master_window.winfo_width()
        height = master_window.winfo_height()
        x_offset = (master_window.winfo_screenwidth() - width) // 2
        y_offset = (master_window.winfo_screenheight() - height) // 2
        master_window.geometry(f"+{x_offset}+{y_offset}")

    def create_form_entry(self, label, variable):
        form_field_container = ttk.Frame(self)
        form_field_container.pack(fill=X, expand=YES, pady=15)

        form_field_label = ttk.Label(master=form_field_container, text=label, width=25)
        form_field_label.pack(side=LEFT, padx=12)

        form_input = ttk.Entry(master=form_field_container, textvariable=variable)
        form_input.pack(side=LEFT, padx=12, fill=X, expand=YES)

        return form_input

    def create_buttonbox(self):
        button_container = ttk.Frame(self)
        button_container.pack(fill=X, expand=YES, pady=(15, 10))

        cancel_btn = ttk.Button(
            master=button_container,
            text="Cancel",
            command=self.on_cancel,
            bootstyle=DANGER,
            width=6,
        )

        cancel_btn.pack(side=RIGHT, padx=5)

        submit_btn = ttk.Button(
            master=button_container,
            text="Submit",
            command=self.on_submit,
            bootstyle=SUCCESS,
            width=6,
        )

        submit_btn.pack(side=RIGHT, padx=5)

    def create_table(self):
        coldata = [
            {"text": "Name"},
            {"text": "Address", "stretch": False},
            {"text": "Doctor Handling"},
            {"text": "Age", "stretch": False},
            {"text": "Queue Number", "stretch": False}  # New column for Queue Number
        ]

        print(self.data_df)

    def on_submit(self):
        name = self.name.get()
        address = self.address.get()
        doctor_handling = self.doctor_handling.get()
        age = self.final_score_input.get()
        
        
        nomor_antrian = self.get_next_queue_number()

        print("Name:", name)
        print("Address: ", address)
        print("Doctor Handling:", doctor_handling)
        print("Age:", age)

        # Increment the queue counter
        self.queue_counter += 1

        # Add data to the DataFrame
        new_data = pd.DataFrame({
            "Name": [name],
            "Address": [address],
            "Doctor Handling": [doctor_handling],
            "Age": [age],
            "Queue Number": [self.queue_counter]
        })

        # Append the new data to the existing DataFrame
        self.data_df = pd.concat([self.data_df, new_data], ignore_index=True)

        # Save the DataFrame to an Excel file, appending to existing file if it exists
        excel_filename = 'hospital_data.xlsx'
        mode = 'w' if not pd.DataFrame(self.data_df).empty else 'a'
        with pd.ExcelWriter(excel_filename, mode=mode, engine='openpyxl') as writer:
            self.data_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=0, header=not bool(writer.sheets))
        print(f"Data saved to {excel_filename}")

        # Display success message with queue number
        success_message = f"Data berhasil disimpan!\nNomer antrian: {self.queue_counter}"
        toast = ToastNotification(
            title="Submission successful!",
            message=success_message,
            duration=3000,
        )
        toast.show_toast()
    def get_next_queue_number(self):
        lokasi_file = Path("Hospital_data.xlsx").expanduser().resolve()

        try:
            workbook = load_workbook(lokasi_file)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet['A1'] = "Nomor Antrian"


        try:
            nomor_antrian = int(sheet['A1'].value)
        except (ValueError, TypeError):
            nomor_antrian = 1

        return nomor_antrian        
        

    def on_cancel(self):
        self.quit()


if __name__ == "__main__":
    app = ttk.Window("AppHospital", "superhero", resizable=(False, False))
    hospital_app = Hospital(app)
    app.mainloop()
